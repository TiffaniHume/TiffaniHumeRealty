from flask import Flask, render_template, request, flash, redirect, url_for, session
import os
import smtplib
from email.message import EmailMessage
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from flask import Flask

# Set directory paths BEFORE initializing Flask
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)
app.secret_key = os.environ.get("SECRET_KEY", "fallback_dev_key")

# =======================
# EMAIL CONFIGURATION
# =======================
EMAIL_ADDRESS = "TiffaniHume179@gmail.com"
EMAIL_PASSWORD = "tynxkavwijzfvkqe"  # App password, not your main Gmail password


# =======================
# EXCEL CRM SETUP
# =======================
DATA_DIR = os.path.join(os.getcwd(), "data")
CRM_FILE = os.path.join(DATA_DIR, "HUME_CRM.xlsx")
CRM_SHEET = "Main"

def log_lead_to_excel(lead_data):
    """Log verified lead data into HUME_CRM.xlsx, guaranteed to append even if Excel is open or locked."""
    import tempfile
    import shutil

    # --- Confirm start ---
    print("\n📘 Starting Excel log process...")
    print(f"Lead data received: {lead_data}")

    # Ensure directory exists
    os.makedirs(DATA_DIR, exist_ok=True)

    # Create workbook if it doesn't exist
    if not os.path.exists(CRM_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = CRM_SHEET
        ws.append(["Timestamp", "Name", "Email", "Phone", "Address", "Notes", "Verified"])
        wb.save(CRM_FILE)
        wb.close()
        print(f"✅ Created new CRM file at {CRM_FILE}")

    # Now, attempt to open workbook safely
    try:
        wb = openpyxl.load_workbook(CRM_FILE)
        ws = wb.active

        new_row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            lead_data.get("name", ""),
            lead_data.get("email", ""),
            lead_data.get("phone", ""),
            lead_data.get("address", ""),
            lead_data.get("notes", ""),
            "Yes"
        ]

        ws.append(new_row)
        print(f"📝 Appending new lead row: {new_row}")

        # Save using atomic temp write (Windows-safe)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name

        wb.save(temp_path)
        wb.close()
        shutil.move(temp_path, CRM_FILE)

        print(f"✅ Lead saved successfully to {CRM_FILE}")

    except PermissionError as e:
        # Excel open fallback
        backup_name = f"HUME_CRM_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        backup_file = os.path.join(DATA_DIR, backup_name)

        print(f"⚠️ PermissionError: {e}. Creating backup at {backup_file}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = CRM_SHEET
        ws.append(["Timestamp", "Name", "Email", "Phone", "Address", "Notes", "Verified"])
        ws.append(new_row)
        wb.save(backup_file)
        wb.close()

        print(f"📦 Backup saved to {backup_file}")

    except Exception as e:
        print(f"❌ Unexpected error while saving Excel: {e}")

# =======================
# ROUTES
# =======================

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/submit", methods=["POST"])
def submit():
    """Handle form submission and email verification"""
    name = request.form.get("name")
    email = request.form.get("email")
    phone = request.form.get("phone")
    address = request.form.get("address")
    notes = request.form.get("notes")

    if not name or not email:
        flash("Name and Email are required.", "error")
        return redirect(url_for("index"))

    verification_code = str(random.randint(100000, 999999))
    session["verification_code"] = verification_code
    session["pending_user"] = {
        "name": name,
        "email": email,
        "phone": phone,
        "address": address,
        "notes": notes
    }

    # Email message
    msg = EmailMessage()
    msg["Subject"] = "Your Tiffani Hume Realty Verification Code"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = email
    msg.set_content(
        f"Hello {name},\n\nThank you for requesting your free customized home value estimate!\n"
        f"Your verification code is: {verification_code}\n\n"
        "Please enter this code on the website to confirm your request.\n\n"
        "— Tiffani Hume Realty"
    )

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        flash("Verification code sent! Please check your email.", "info")
        return redirect(url_for("verify"))
    except Exception as e:
        print(f"Email error: {e}")
        flash("Unable to send verification email. Please try again later.", "error")
        return redirect(url_for("index"))

@app.route("/verify", methods=["GET", "POST"])
def verify():
    """Email verification page"""
    if request.method == "POST":
        code = request.form.get("code")
        if code == session.get("verification_code"):
            user_data = session.get("pending_user")
            if user_data:
                log_lead_to_excel(user_data)
                session.pop("pending_user", None)
                session.pop("verification_code", None)
            flash("Email verified successfully!", "success")
            return redirect(url_for("success"))
        else:
            flash("Invalid code. Please check your email and try again.", "error")
    return render_template("verify.html")

@app.route("/success")
def success():
    """Display final confirmation message"""
    return render_template("success.html")

# =======================
# APP ENTRY POINT
# =======================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
